#!/usr/bin/env python3
"""QA 시트(자유 텍스트) → testcases 키워드 형식 '초안' 변환기.

기계적으로 변환 가능한 부분만 자동 채우고, 사람이 검토/보정하도록 초안을 만든다.
- 각 케이스: launch 로 시작
- 절차 문구에서 흔한 탭 동작을 키워드로 매핑(휴리스틱)
- 예상결과의 "..." 인용 문구를 assert_visible 로 추출
- 센서/카운트다운/권한변경/결과/공유 등은 SKIP(사유) 로 표시 → 자동화 대상 아님

사용: python3 tools/qa_to_testcases.py "<QA.xlsx>" [시트명] [출력.xlsx]
"""
import sys, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

TAP_MAP = [  # 절차에 이 문구가 있으면 → (action, target)
    ("러닝 탭",       ("tap_text","러닝")),
    ("시작하기",      ("tap_text","시작하기")),
    ("실내 러닝",     ("tap_text","실내 러닝")),
    ("야외 러닝",     ("tap_text","야외 러닝")),
    ("확인 버튼",     ("tap_text","확인")),
    ("닫기(X)",      ("back","")),
    ("닫기",          ("back","")),
]
# 케이스가 이 단어를 포함하면 자동화 불가로 표시
SKIP_KW = ["케이던스","걸음","카운트다운","일시정지","자동 일시정지","재개","정지(",
           "롱프레스","FINISH","결과 화면","인증샷","공유","권한 미허용","비허용",
           "강제종료","백그라운드","음성","심박","고도","템플릿","토스트","롱프레스"]

def convert(src, sheet, out):
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    # 헤더 위치
    hi = next((i for i,r in enumerate(rows[:6])
               if any((c or "")=="No." for c in r)), 1)
    out_wb = openpyxl.Workbook(); o = out_wb.active; o.title="cases"
    o.append(["TC","이름","action","target","비고"])
    for c in o[1]:
        c.fill=PatternFill("solid",fgColor="1F2937"); c.font=Font(color="FFFFFF",bold=True)
    for r in rows[hi+1:]:
        def g(j): return "" if j>=len(r) or r[j] is None else str(r[j]).replace("\n"," ").strip()
        no,tc,proc,exp = g(0),g(3),g(6),g(7)
        if not tc: continue
        tcid = "TC"+no.replace(".0","")
        skip = next((k for k in SKIP_KW if k in tc or k in exp), None)
        if skip:
            o.append([tcid, tc[:40], "SKIP", "", f"자동화 불가(±): '{skip}' 관련 — 센서/실주행/육안/권한 필요"])
            continue
        o.append([tcid, tc[:40], "launch", "", ""])
        # 절차 → 탭 키워드
        for key,(act,tgt) in TAP_MAP:
            if key in proc:
                o.append([tcid,"",act,tgt, f"절차: {key}"])
        # 예상결과 "..." → assert_visible
        quotes = re.findall(r'"([^"]+)"', exp)
        for q in quotes[:6]:
            o.append([tcid,"","assert_visible", q, "예상결과 인용"])
        if not quotes:
            o.append([tcid,"","TODO","", "예상결과에 인용문구 없음 → 검증 대상 직접 입력 필요"])
    for i,w in enumerate([9,40,15,40,45],1):
        o.column_dimensions[chr(64+i)].width=w
    for row in o.iter_rows(min_row=2):
        for c in row: c.alignment=Alignment(vertical="top",wrap_text=True)
    o.freeze_panes="A2"
    out_wb.save(out)
    print("생성:", out, " 시트:", sheet, " 행:", o.max_row-1)

if __name__=="__main__":
    src = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv)>2 else "모바일_실내러닝"
    out = sys.argv[3] if len(sys.argv)>3 else "data/testcases_from_qa.xlsx"
    convert(src, sheet, out)
