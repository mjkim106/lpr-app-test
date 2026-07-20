"""대시보드용 testcases.xlsx 읽기/쓰기 (비고 포함, TC별 스텝 묶음).

load() → [{"tc":..,"name":..,"steps":[{"action","target","note"}, ...]}, ...]
save(cases) → cases 시트를 재작성(다른 시트는 보존).
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "data", "testcases.xlsx")
HEADER = ["TC", "이름", "action", "target", "비고"]


def load(path=XLSX):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["cases"]
    order, cases = [], {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        def g(j):
            return "" if j >= len(row) or row[j] is None else str(row[j]).strip()
        tc, name, action, target, note = g(0), g(1), g(2), g(3), g(4)
        if not tc or not action:
            continue
        if tc not in cases:
            cases[tc] = {"tc": tc, "name": "", "steps": []}
            order.append(tc)
        if name:
            cases[tc]["name"] = name
        cases[tc]["steps"].append({"action": action, "target": target, "note": note})
    return [cases[tc] for tc in order]


def save(cases, path=XLSX):
    # 기존 워크북 로드(다른 시트 보존), cases 시트만 재작성
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        if "cases" in wb.sheetnames:
            wb.remove(wb["cases"])
        ws = wb.create_sheet("cases", 0)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "cases"
    ws.append(HEADER)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1F2937")
        c.font = Font(color="FFFFFF", bold=True)
    for case in cases:
        tc = case["tc"]
        steps = case.get("steps", [])
        for idx, st in enumerate(steps):
            name = case.get("name", "") if idx == 0 else ""
            ws.append([tc, name, st.get("action", ""), st.get("target", ""), st.get("note", "")])
    for i, w in enumerate([12, 26, 18, 26, 34], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(path)
